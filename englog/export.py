"""Export session data as narrative summary + xlsx timesheet.

Produces two outputs:
1. AI narrative summary (markdown) — what was done and why
2. Structured timesheet (xlsx) — 30-minute window blocks with task title + description

Key design decisions:
- Time is divided into fixed 30-min windows (09:00-09:30, 09:30-10:00, etc.)
- A session always maps onto these windows — never creates arbitrary time ranges
- The first and last window may be partial (e.g., session starts at 09:12, first window is 09:00-09:30 with 18 min active)
- Each window gets ONE primary task (the dominant activity in that window)
- The AI infers the task from notes + active window/process metadata
"""

import json
import re
from datetime import datetime, timedelta
from math import ceil
from pathlib import Path
from typing import Optional

import requests

from englog.config import OLLAMA_BASE_URL, OLLAMA_MODEL, OLLAMA_NUM_CTX, MAX_CONTEXT_CHARS, OLLAMA_TIMEOUT, DATA_DIR
from englog import database as db

EXPORTS_DIR = DATA_DIR / "exports"


def _get_session_time_range(session: dict) -> tuple[datetime, datetime]:
    """Parse session start/end into datetime objects."""
    fmt = "%Y-%m-%d %H:%M:%S"
    start = datetime.strptime(session["started_at"], fmt)
    end_str = session.get("ended_at") or datetime.now().strftime(fmt)
    end = datetime.strptime(end_str, fmt)
    return start, end


def _round_down_30(dt: datetime) -> datetime:
    """Round a datetime down to the nearest 30-min boundary."""
    return dt.replace(minute=(dt.minute // 30) * 30, second=0, microsecond=0)


def _round_up_30(dt: datetime) -> datetime:
    """Round a datetime up to the nearest 30-min boundary."""
    if dt.minute % 30 == 0 and dt.second == 0:
        return dt.replace(second=0, microsecond=0)
    base = dt.replace(minute=(dt.minute // 30) * 30, second=0, microsecond=0)
    return base + timedelta(minutes=30)


def _compute_slots(session_start: datetime, session_end: datetime) -> list[dict]:
    """Generate the fixed 30-min window grid with actual active durations.

    For a session 09:12 to 17:45:
    - First slot: 09:00-09:30 (18 min active: 09:12 to 09:30)
    - Middle slots: 09:30-10:00 ... 17:00-17:30 (30 min each)
    - Last slot: 17:30-18:00 (15 min active: 17:30 to 17:45)

    For a short session 01:49 to 01:52 (3 min):
    - Single slot: 01:30-02:00 (3 min active)
    """
    window_start = _round_down_30(session_start)
    window_end = _round_up_30(session_end)

    # Edge case: entire session fits in one window
    if window_start == _round_down_30(session_end):
        duration = max(int((session_end - session_start).total_seconds() / 60), 1)
        return [{
            "window_start": window_start.strftime("%H:%M"),
            "window_end": (window_start + timedelta(minutes=30)).strftime("%H:%M"),
            "active_minutes": duration,
            "task_title": "",
            "task_description": "",
        }]

    slots = []
    current = window_start
    while current < window_end:
        next_boundary = current + timedelta(minutes=30)

        # Calculate active minutes within this window
        active_start = max(current, session_start)
        active_end = min(next_boundary, session_end)
        active_minutes = max(int((active_end - active_start).total_seconds() / 60), 0)

        if active_minutes > 0:
            slots.append({
                "window_start": current.strftime("%H:%M"),
                "window_end": next_boundary.strftime("%H:%M"),
                "active_minutes": active_minutes,
                "task_title": "",
                "task_description": "",
            })

        current = next_boundary

    return slots


TIMESHEET_SYSTEM_PROMPT = """You are EngLog, an engineering logbook assistant. You receive a session 
timeline (notes + active window metadata) and a pre-computed list of 30-minute time windows.

Your job: for each window, determine the PRIMARY activity that happened during that window.

Rules:
- Each window gets exactly ONE task_title (2-5 words) and ONE task_description (1 sentence)
- The task should represent the MAIN activity in that window — if multiple things happened, pick the dominant one
- Use active window/process data to infer what was happening between notes:
  * EXCEL.EXE -> spreadsheet work (mention the file name from window title if visible)
  * chrome.exe + Gmail -> email correspondence  
  * chrome.exe + other -> web research / browsing (mention the topic if visible)
  * code.exe / pycharm / VS Code -> coding (mention file/project if visible)
  * explorer.exe -> file management
- Combine with nearby notes for richer descriptions
- If a window has no captures and no notes nearby, use "Break / Away" as the task
- Be specific: "Updated CMG mass budget spreadsheet" not "Worked on spreadsheet"

CRITICAL: You must return EXACTLY the windows provided — same start/end times, same count. 
Do NOT invent new windows or change the times.

Respond with ONLY valid JSON, no markdown, no backticks. Format:
{
  "blocks": [
    {
      "window_start": "09:00",
      "window_end": "09:30",
      "task_title": "Literature review",
      "task_description": "Reviewed ECSS-E-ST-60 standard for CMG sizing methodology in Chrome"
    }
  ]
}
"""


def _build_timesheet_context(session_id: int) -> tuple[str, list[dict]]:
    """Build context for the AI and return (context_text, pre_computed_slots)."""
    session = db.get_session(session_id)
    if not session:
        return "", []

    notes = db.get_session_notes(session_id)
    captures = db.get_session_captures(session_id)
    start, end = _get_session_time_range(session)
    slots = _compute_slots(start, end)

    # Fetch rich project context, rules, and examples if available
    project = db.get_project(session['project_name'])
    project_context = (project.get("context") or "").strip() if project else ""
    project_rules = (project.get("rules") or "").strip() if project else ""

    lines = []
    lines.append(f"Project: {session['project_name']}")
    if project_context:
        lines.append(f"Project context: {project_context}")
    if project_rules:
        lines.append(f"Project-specific rules (follow these when classifying tasks): {project_rules}")
    lines.append(f"Session: {session['started_at']} to {session.get('ended_at', 'ongoing')}")
    lines.append(f"Total duration: {int((end - start).total_seconds() / 60)} minutes")
    lines.append("")

    # Tell the AI exactly which windows to fill
    lines.append("=== WINDOWS TO FILL ===")
    for slot in slots:
        lines.append(f"  {slot['window_start']}-{slot['window_end']} ({slot['active_minutes']} min active)")
    lines.append("")

    # Merge notes and captures into timeline
    events = []
    for note in notes:
        events.append({"ts": note["timestamp"], "type": "NOTE", "detail": f"({note['note_type']}) {note['content']}"})

    last_key = None
    for cap in captures:
        key = f"{cap.get('active_process', '')}|{cap.get('active_window', '')}"
        if key != last_key:
            events.append({
                "ts": cap["timestamp"],
                "type": "CONTEXT",
                "detail": f"{cap.get('active_process', '?')} — \"{cap.get('active_window', '?')}\"",
            })
            last_key = key

    events.sort(key=lambda e: e["ts"])

    # Safety cap: trim capture events if context would exceed budget
    header_chars = len("\n".join(lines))
    budget = MAX_CONTEXT_CHARS - header_chars
    note_events = [e for e in events if e["type"] == "NOTE"]
    ctx_events = [e for e in events if e["type"] == "CONTEXT"]
    estimated_chars = len(events) * 80
    if estimated_chars > budget and len(ctx_events) > 10:
        notes_chars = len(note_events) * 80
        max_ctx = max((budget - notes_chars) // 80, 10)
        if len(ctx_events) > max_ctx:
            step = (len(ctx_events) - 1) / (max_ctx - 1)
            indices = {round(i * step) for i in range(max_ctx)}
            ctx_events = [c for i, c in enumerate(ctx_events) if i in indices]
        events = note_events + ctx_events
        events.sort(key=lambda e: e["ts"])

    lines.append("=== TIMELINE ===")
    for ev in events:
        lines.append(f"[{ev['ts']}] {ev['type']}: {ev['detail']}")

    return "\n".join(lines), slots


def generate_timesheet_data(session_id: int) -> list[dict]:
    """Ask the LLM to fill in the pre-computed 30-min windows."""
    context, slots = _build_timesheet_context(session_id)
    if not context or not slots:
        return []

    payload = {
        "model": OLLAMA_MODEL,
        "system": TIMESHEET_SYSTEM_PROMPT,
        "prompt": f"Generate the timesheet for this session:\n\n{context}",
        "stream": False,
        "options": {"temperature": 0.2, "num_predict": 4096, "num_ctx": OLLAMA_NUM_CTX},
        "format": "json",
    }

    try:
        resp = requests.post(f"{OLLAMA_BASE_URL}/api/generate", json=payload, timeout=OLLAMA_TIMEOUT)
        resp.raise_for_status()
        raw = resp.json().get("response", "").strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

        data = json.loads(raw)
        ai_blocks = data.get("blocks", [])

        # Merge AI output back onto our pre-computed slots
        # The AI might reorder, skip, or use slightly different keys
        ai_lookup = {}
        for b in ai_blocks:
            key = b.get("window_start", b.get("start", ""))
            ai_lookup[key] = b

        for slot in slots:
            ai_data = ai_lookup.get(slot["window_start"], {})
            slot["task_title"] = ai_data.get("task_title", "Unclassified activity")
            slot["task_description"] = ai_data.get("task_description", "")

        return slots

    except (json.JSONDecodeError, requests.exceptions.ReadTimeout, requests.RequestException) as e:
        import sys
        print(
            f"[englog] Timesheet AI classification failed: {e}\n"
            f"[englog] The time grid and notes will still be exported. "
            f"Re-run with: englog export {session_id}",
            file=sys.stderr,
        )
        # Return slots with placeholder text so the xlsx still generates
        for slot in slots:
            slot["task_title"] = "(Fill manually)"
            slot["task_description"] = "AI classification failed — re-run 'englog export' to retry, or fill in manually"
        return slots


def export_xlsx(session_id: int, output_path: Optional[str] = None, ollama_available: bool = True) -> str:
    """Export a session timesheet as a formatted .xlsx file.

    Always produces an xlsx — if Ollama is unavailable, the time grid and notes
    are still exported with placeholder task descriptions.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    session = db.get_session(session_id)
    if not session:
        raise ValueError(f"Session {session_id} not found")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if not output_path:
        date_str = session["started_at"][:10].replace("-", "")
        output_path = str(EXPORTS_DIR / f"englog_{session['project_name']}_{date_str}_s{session_id}.xlsx")

    if ollama_available:
        blocks = generate_timesheet_data(session_id)
    else:
        # Generate the time grid without AI — still useful as a structural skeleton
        start, end = _get_session_time_range(session)
        blocks = _compute_slots(start, end)
        for slot in blocks:
            slot["task_title"] = "(Fill manually)"
            slot["task_description"] = "AI unavailable — add task description manually"
    notes = db.get_session_notes(session_id)

    wb = Workbook()

    # ── Sheet 1: Timesheet ──────────────────────────────
    ws = wb.active
    ws.title = "Timesheet"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Arial", size=10)
    time_font = Font(name="Arial", size=10, bold=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"].value = f"EngLog Timesheet — {session['project_name']}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle with total time
    start, end = _get_session_time_range(session)
    total_minutes = int((end - start).total_seconds() / 60)
    hours, mins = divmod(total_minutes, 60)
    ws.merge_cells("A2:D2")
    ws["A2"].value = (
        f"Session #{session_id} | {session['started_at']} → {session.get('ended_at', 'ongoing')}"
        f" | Total: {hours}h {mins}m"
    )
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["Time Window", "Active", "Task", "Description"]
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 60

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 25

    # Data rows — duration is now calculated, not hardcoded
    for i, block in enumerate(blocks):
        row = 5 + i
        time_slot = f"{block['window_start']} – {block['window_end']}"
        active = block["active_minutes"]
        duration_str = f"{active} min"
        task = block.get("task_title", "")
        desc = block.get("task_description", "")

        ws.cell(row=row, column=1, value=time_slot).font = time_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=duration_str).font = cell_font
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=3, value=task).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=3).alignment = wrap_align
        ws.cell(row=row, column=4, value=desc).font = cell_font
        ws.cell(row=row, column=4).alignment = wrap_align
        ws.row_dimensions[row].height = 30

        if i % 2 == 1:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = alt_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

    # ── Sheet 2: Notes ──────────────────────────────────
    ws2 = wb.create_sheet("Notes")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 70

    for col_idx, header in enumerate(["Timestamp", "Type", "Content"], 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    type_colors = {
        "decision": PatternFill("solid", fgColor="E8D5F5"),
        "blocker": PatternFill("solid", fgColor="FCDEDE"),
        "observation": PatternFill("solid", fgColor="DEEAF6"),
    }

    for i, n in enumerate(notes):
        row = 2 + i
        ws2.cell(row=row, column=1, value=n["timestamp"]).font = cell_font
        type_cell = ws2.cell(row=row, column=2, value=n["note_type"].upper())
        type_cell.font = Font(name="Arial", size=10, bold=True)
        type_cell.fill = type_colors.get(n["note_type"], PatternFill())
        type_cell.alignment = center_align
        ws2.cell(row=row, column=3, value=n["content"]).font = cell_font
        ws2.cell(row=row, column=3).alignment = wrap_align
        for col in range(1, 4):
            ws2.cell(row=row, column=col).border = thin_border

    # ── Sheet 3: Summary ────────────────────────────────
    if session.get("summary"):
        ws3 = wb.create_sheet("AI Summary")
        ws3.column_dimensions["A"].width = 100
        ws3["A1"].value = "AI-Generated Session Summary"
        ws3["A1"].font = Font(name="Arial", bold=True, size=12, color="2F5496")
        for i, line in enumerate(session["summary"].split("\n")):
            ws3.cell(row=3 + i, column=1, value=line).font = cell_font

    wb.save(output_path)
    return output_path
